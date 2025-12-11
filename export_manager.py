#!/usr/bin/env python3
"""
Export Manager Module
Handles exporting results to various formats (CSV, Excel, JSON, HTML for printing)
"""

import csv
import json
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path
import html


class ExportManager:
    """Manages exporting search results to various formats"""
    
    def __init__(self, output_dir: str = None):
        """
        Initialize the export manager
        
        Args:
            output_dir: Default directory for exports
        """
        if output_dir is None:
            output_dir = Path.home() / 'Downloads' / 'FinnDealFinder'
        
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
    def export(
        self,
        items: List[Dict],
        format_type: str = 'csv',
        filename: str = None,
        include_analysis: bool = True
    ) -> str:
        """
        Export items to specified format
        
        Args:
            items: List of items to export
            format_type: Export format ('csv', 'excel', 'json')
            filename: Custom filename (optional)
            include_analysis: Include deal analysis data
            
        Returns:
            Path to exported file
        """
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'finn_deals_{timestamp}'
        
        if format_type == 'csv':
            return self._export_csv(items, filename, include_analysis)
        elif format_type == 'excel':
            return self._export_excel(items, filename, include_analysis)
        elif format_type == 'json':
            return self._export_json(items, filename, include_analysis)
        else:
            raise ValueError(f"Unsupported format: {format_type}")
            
    def _export_csv(
        self,
        items: List[Dict],
        filename: str,
        include_analysis: bool
    ) -> str:
        """Export items to CSV format"""
        filepath = self.output_dir / f'{filename}.csv'
        
        # Define columns
        columns = [
            'Title',
            'Price (NOK)',
            'Average Price (NOK)',
            'Deal Score (%)',
            'Recommendation',
            'Location',
            'Condition',
            'Posted',
            'Seller Type',
            'URL',
            'FINN ID'
        ]
        
        if include_analysis:
            columns.extend([
                'Price vs Average',
                'Price Factor',
                'Condition Factor',
            ])
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            
            for item in items:
                row = {
                    'Title': item.get('title', ''),
                    'Price (NOK)': item.get('price', ''),
                    'Average Price (NOK)': item.get('avg_price', ''),
                    'Deal Score (%)': item.get('deal_score', ''),
                    'Recommendation': item.get('recommendation', ''),
                    'Location': item.get('location', ''),
                    'Condition': item.get('condition', ''),
                    'Posted': item.get('posted', ''),
                    'Seller Type': item.get('seller_type', ''),
                    'URL': item.get('url', ''),
                    'FINN ID': item.get('id', '')
                }
                
                if include_analysis:
                    factors = item.get('deal_factors', {})
                    details = factors.get('details', {})
                    
                    row.update({
                        'Price vs Average': details.get('price_vs_avg', ''),
                        'Price Factor': factors.get('price_factor', ''),
                        'Condition Factor': factors.get('condition_factor', ''),
                    })
                
                writer.writerow(row)
        
        return str(filepath)
        
    def _export_excel(
        self,
        items: List[Dict],
        filename: str,
        include_analysis: bool
    ) -> str:
        """Export items to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback to CSV if openpyxl not available
            print("openpyxl not available, falling back to CSV")
            return self._export_csv(items, filename, include_analysis)
        
        filepath = self.output_dir / f'{filename}.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FINN Deals"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        deal_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        great_deal_fill = PatternFill(start_color="86EFAC", end_color="86EFAC", fill_type="solid")
        excellent_deal_fill = PatternFill(start_color="4ADE80", end_color="4ADE80", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            'Title',
            'Price (NOK)',
            'Avg Price (NOK)',
            'Deal Score',
            'Recommendation',
            'Location',
            'Condition',
            'Posted',
            'Seller',
            'URL'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data rows
        for row_num, item in enumerate(items, 2):
            deal_score = item.get('deal_score', 0)
            
            # Determine row fill based on deal score
            if deal_score >= 90:
                row_fill = excellent_deal_fill
            elif deal_score >= 80:
                row_fill = great_deal_fill
            elif deal_score >= 70:
                row_fill = deal_fill
            else:
                row_fill = None
            
            row_data = [
                item.get('title', ''),
                item.get('price', 0),
                item.get('avg_price', 0),
                f"{deal_score}%",
                item.get('recommendation', ''),
                item.get('location', ''),
                item.get('condition', ''),
                item.get('posted', ''),
                item.get('seller_type', ''),
                item.get('url', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = max(
                len(str(ws.cell(row=row, column=col).value or ''))
                for row in range(1, min(len(items) + 2, 100))  # Check first 100 rows
            )
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add statistics sheet
        stats_ws = wb.create_sheet("Statistics")
        
        if items:
            prices = [item.get('price', 0) for item in items if item.get('price', 0) > 0]
            deal_scores = [item.get('deal_score', 0) for item in items]
            hot_deals = len([s for s in deal_scores if s >= 70])
            
            stats = [
                ('Statistic', 'Value'),
                ('Total Items', len(items)),
                ('Hot Deals (70%+)', hot_deals),
                ('Average Price', f"{sum(prices)/len(prices):,.0f} kr" if prices else "N/A"),
                ('Min Price', f"{min(prices):,} kr" if prices else "N/A"),
                ('Max Price', f"{max(prices):,} kr" if prices else "N/A"),
                ('Average Deal Score', f"{sum(deal_scores)/len(deal_scores):.1f}%" if deal_scores else "N/A"),
                ('Best Deal Score', f"{max(deal_scores)}%" if deal_scores else "N/A"),
                ('Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ]
            
            for row_num, (label, value) in enumerate(stats, 1):
                stats_ws.cell(row=row_num, column=1, value=label)
                stats_ws.cell(row=row_num, column=2, value=value)
                
                if row_num == 1:
                    stats_ws.cell(row=row_num, column=1).font = header_font
                    stats_ws.cell(row=row_num, column=1).fill = header_fill
                    stats_ws.cell(row=row_num, column=2).font = header_font
                    stats_ws.cell(row=row_num, column=2).fill = header_fill
        
        wb.save(filepath)
        return str(filepath)
        
    def _export_json(
        self,
        items: List[Dict],
        filename: str,
        include_analysis: bool
    ) -> str:
        """Export items to JSON format"""
        filepath = self.output_dir / f'{filename}.json'
        
        # Prepare export data
        export_data = {
            'export_info': {
                'generated_at': datetime.now().isoformat(),
                'total_items': len(items),
                'source': 'FINN.no Deal Finder Pro'
            },
            'statistics': self._calculate_export_stats(items),
            'items': items if include_analysis else self._strip_analysis(items)
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        return str(filepath)
        
    def _strip_analysis(self, items: List[Dict]) -> List[Dict]:
        """Remove analysis data from items for basic export"""
        stripped = []
        for item in items:
            stripped_item = {
                'id': item.get('id'),
                'title': item.get('title'),
                'price': item.get('price'),
                'location': item.get('location'),
                'condition': item.get('condition'),
                'posted': item.get('posted'),
                'url': item.get('url'),
                'seller_type': item.get('seller_type')
            }
            stripped.append(stripped_item)
        return stripped
        
    def _calculate_export_stats(self, items: List[Dict]) -> Dict:
        """Calculate statistics for export"""
        if not items:
            return {}
        
        prices = [item.get('price', 0) for item in items if item.get('price', 0) > 0]
        deal_scores = [item.get('deal_score', 0) for item in items]
        
        return {
            'total_items': len(items),
            'items_with_price': len(prices),
            'average_price': sum(prices) / len(prices) if prices else 0,
            'min_price': min(prices) if prices else 0,
            'max_price': max(prices) if prices else 0,
            'average_deal_score': sum(deal_scores) / len(deal_scores) if deal_scores else 0,
            'best_deal_score': max(deal_scores) if deal_scores else 0,
            'excellent_deals': len([s for s in deal_scores if s >= 90]),
            'great_deals': len([s for s in deal_scores if 80 <= s < 90]),
            'good_deals': len([s for s in deal_scores if 70 <= s < 80]),
        }
        
    def create_print_report(
        self,
        items: List[Dict],
        search_params: Dict = None,
        title: str = "FINN.no Deal Report"
    ) -> str:
        """Create an HTML report for printing"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filepath = self.output_dir / f'finn_report_{timestamp}.html'
        
        # Calculate statistics
        stats = self._calculate_export_stats(items)
        
        # Sort items by deal score
        sorted_items = sorted(items, key=lambda x: x.get('deal_score', 0), reverse=True)
        
        # Generate HTML
        html_content = self._generate_print_html(
            title=title,
            items=sorted_items,
            stats=stats,
            search_params=search_params
        )
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return str(filepath)
        
    def _generate_print_html(
        self,
        title: str,
        items: List[Dict],
        stats: Dict,
        search_params: Dict = None
    ) -> str:
        """Generate HTML content for print report"""
        
        # Generate item rows
        item_rows = []
        for i, item in enumerate(items, 1):
            deal_score = item.get('deal_score', 0)
            
            # Determine deal class
            if deal_score >= 90:
                deal_class = 'excellent'
                deal_badge = '🔥 EXCELLENT'
            elif deal_score >= 80:
                deal_class = 'great'
                deal_badge = '⭐ GREAT'
            elif deal_score >= 70:
                deal_class = 'good'
                deal_badge = '👍 GOOD'
            else:
                deal_class = 'normal'
                deal_badge = ''
            
            price = item.get('price', 0)
            avg_price = item.get('avg_price', 0)
            
            savings = ''
            if avg_price > price > 0:
                savings = f'<span class="savings">💰 Save {avg_price - price:,.0f} kr</span>'
            
            item_rows.append(f'''
                <tr class="item-row {deal_class}">
                    <td class="rank">{i}</td>
                    <td class="title">
                        <strong>{html.escape(item.get('title', 'N/A')[:60])}</strong>
                        {f'<span class="deal-badge">{deal_badge}</span>' if deal_badge else ''}
                        <br><small class="location">📍 {html.escape(item.get('location', 'N/A'))}</small>
                    </td>
                    <td class="price">
                        <strong>{price:,.0f} kr</strong>
                        {savings}
                    </td>
                    <td class="score">
                        <div class="score-bar">
                            <div class="score-fill {deal_class}" style="width: {deal_score}%"></div>
                        </div>
                        <span>{deal_score}%</span>
                    </td>
                    <td class="condition">{html.escape(item.get('condition', 'N/A'))}</td>
                    <td class="link">
                        <a href="{item.get('url', '#')}" target="_blank">View</a>
                    </td>
                </tr>
            ''')
        
        # Build search info
        search_info = ''
        if search_params:
            search_details = []
            if search_params.get('keyword'):
                search_details.append(f"🔍 <strong>{html.escape(search_params['keyword'])}</strong>")
            if search_params.get('category'):
                search_details.append(f"📁 {html.escape(search_params['category'])}")
            if search_params.get('location'):
                search_details.append(f"📍 {html.escape(search_params['location'])}")
            
            if search_details:
                search_info = ' | '.join(search_details)
        
        html_content = f'''
<!DOCTYPE html>
<html lang="no">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{html.escape(title)}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
            line-height: 1.6;
            color: #1e293b;
            background: #f8fafc;
            padding: 20px;
        }}
        
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
            overflow: hidden;
        }}
        
        .header {{
            background: linear-gradient(135deg, #4f46e5 0%, #7c3aed 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        
        .header h1 {{
            font-size: 2rem;
            margin-bottom: 10px;
        }}
        
        .header .subtitle {{
            opacity: 0.9;
            font-size: 1.1rem;
        }}
        
        .search-info {{
            background: rgba(255,255,255,0.1);
            padding: 10px 20px;
            border-radius: 8px;
            margin-top: 15px;
            display: inline-block;
        }}
        
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 15px;
            padding: 25px;
            background: #f1f5f9;
        }}
        
        .stat-card {{
            background: white;
            padding: 15px;
            border-radius: 8px;
            text-align: center;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }}
        
        .stat-value {{
            font-size: 1.8rem;
            font-weight: bold;
            color: #4f46e5;
        }}
        
        .stat-label {{
            font-size: 0.85rem;
            color: #64748b;
            margin-top: 5px;
        }}
        
        .deals-table {{
            width: 100%;
            border-collapse: collapse;
        }}
        
        .deals-table th {{
            background: #1e293b;
            color: white;
            padding: 15px 12px;
            text-align: left;
            font-weight: 600;
            position: sticky;
            top: 0;
        }}
        
        .item-row {{
            border-bottom: 1px solid #e2e8f0;
            transition: background 0.2s;
        }}
        
        .item-row:hover {{
            background: #f8fafc;
        }}
        
        .item-row td {{
            padding: 15px 12px;
            vertical-align: top;
        }}
        
        .item-row.excellent {{
            background: linear-gradient(90deg, #dcfce7 0%, white 100%);
        }}
        
        .item-row.great {{
            background: linear-gradient(90deg, #fef9c3 0%, white 100%);
        }}
        
        .item-row.good {{
            background: linear-gradient(90deg, #e0f2fe 0%, white 100%);
        }}
        
        .rank {{
            width: 40px;
            font-weight: bold;
            color: #64748b;
        }}
        
        .title {{
            max-width: 350px;
        }}
        
        .title strong {{
            color: #1e293b;
        }}
        
        .location {{
            color: #64748b;
        }}
        
        .deal-badge {{
            display: inline-block;
            padding: 2px 8px;
            border-radius: 4px;
            font-size: 0.75rem;
            font-weight: bold;
            margin-left: 8px;
            background: #4f46e5;
            color: white;
        }}
        
        .price {{
            font-size: 1.1rem;
        }}
        
        .savings {{
            display: block;
            color: #059669;
            font-size: 0.85rem;
            margin-top: 4px;
        }}
        
        .score {{
            width: 120px;
        }}
        
        .score-bar {{
            width: 80px;
            height: 8px;
            background: #e2e8f0;
            border-radius: 4px;
            overflow: hidden;
            display: inline-block;
            vertical-align: middle;
            margin-right: 8px;
        }}
        
        .score-fill {{
            height: 100%;
            border-radius: 4px;
            transition: width 0.3s;
        }}
        
        .score-fill.excellent {{
            background: #22c55e;
        }}
        
        .score-fill.great {{
            background: #84cc16;
        }}
        
        .score-fill.good {{
            background: #eab308;
        }}
        
        .score-fill.normal {{
            background: #94a3b8;
        }}
        
        .link a {{
            color: #4f46e5;
            text-decoration: none;
            font-weight: 500;
        }}
        
        .link a:hover {{
            text-decoration: underline;
        }}
        
        .footer {{
            text-align: center;
            padding: 20px;
            background: #f1f5f9;
            color: #64748b;
            font-size: 0.9rem;
        }}
        
        @media print {{
            body {{
                padding: 0;
                background: white;
            }}
            
            .container {{
                box-shadow: none;
            }}
            
            .header {{
                background: #4f46e5 !important;
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }}
            
            .item-row.excellent,
            .item-row.great,
            .item-row.good {{
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }}
            
            .link {{
                display: none;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🏆 {html.escape(title)}</h1>
            <p class="subtitle">Generated by FINN.no Deal Finder Pro</p>
            {f'<div class="search-info">{search_info}</div>' if search_info else ''}
        </div>
        
        <div class="stats-grid">
            <div class="stat-card">
                <div class="stat-value">{stats.get('total_items', 0)}</div>
                <div class="stat-label">Total Items</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats.get('excellent_deals', 0) + stats.get('great_deals', 0) + stats.get('good_deals', 0)}</div>
                <div class="stat-label">Hot Deals</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats.get('average_price', 0):,.0f} kr</div>
                <div class="stat-label">Avg. Price</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats.get('best_deal_score', 0)}%</div>
                <div class="stat-label">Best Score</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats.get('min_price', 0):,.0f} kr</div>
                <div class="stat-label">Min Price</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{stats.get('max_price', 0):,.0f} kr</div>
                <div class="stat-label">Max Price</div>
            </div>
        </div>
        
        <table class="deals-table">
            <thead>
                <tr>
                    <th>#</th>
                    <th>Item</th>
                    <th>Price</th>
                    <th>Deal Score</th>
                    <th>Condition</th>
                    <th>Link</th>
                </tr>
            </thead>
            <tbody>
                {''.join(item_rows)}
            </tbody>
        </table>
        
        <div class="footer">
            <p>Report generated on {datetime.now().strftime('%B %d, %Y at %H:%M')}</p>
            <p>Data source: FINN.no | Tool: Deal Finder Pro v1.0</p>
        </div>
    </div>
</body>
</html>
'''
        
        return html_content
